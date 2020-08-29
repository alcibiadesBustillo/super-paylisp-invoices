import openpyxl as xl
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from datetime import datetime

from reportlab.pdfbase.pdfmetrics import stringWidth
from reportlab.pdfbase.ttfonts import TTFont


pdfmetrics.registerFont(TTFont('Vera', 'Vera.ttf'))

work_book = xl.load_workbook('data.xlsx')
sheet = work_book['employees']

print(sheet.cell(2,3).value)

# settings of pdf
page_width = 2156
page_height = 3050
spread = 100
start = 200
start_2 = 900

date = datetime.now()
company_name = "@lbuz@r"

def create_paylisp():
    for i in range(2,3):
        emp_id = sheet.cell(row=i, column=1).value #Done
        emp_name = sheet.cell(row=i, column=2).value #Done
        emp_surname = sheet.cell(row=i, column=3).value #Done
        emp_gross_salary = sheet.cell(row=i, column=4).value
        emp_pension_contribution = sheet.cell(row=i, column=5).value
        emp_health_insurance = sheet.cell(row=i, column=6).value
        emp_personal_income_tax = sheet.cell(row=i, column=7).value
        emp_bonus_payment = sheet.cell(row=i, column=8).value
        emp_deduction = sheet.cell(row=i, column=9).value
        emp_net_salary = sheet.cell(row=i, column=10).value
        emp_email = sheet.cell(row=i, column=11).value
        emp_phone_number = sheet.cell(row=i, column=12).value

        c = canvas.Canvas("{}_{}_{}_{}.pdf".format(emp_name, emp_surname, emp_id, date.strftime("%y_%m_%d")))
        c.setPageSize((page_width, page_height))
        
        c.setFont('Vera', 80)
        text_width = stringWidth(company_name, 'Vera', 80)
        c.drawString((page_width-text_width)/2, 2900, company_name)

        text = 'Salary calculation for period {}'.format(date.strftime("%B %Y"))
        text_width = stringWidth(text, 'Vera', 55)
        c.setFont('Vera', 55)
        c.drawString((page_width-text_width)/2, 2700, text)
        
        y = 2500
        c.setFont('Vera', 45)
        c.drawString(start, y, 'Employee\'s id:')
        c.drawString(start_2, y, str(emp_id))
        y -= spread

        c.drawString(start, y, 'Employee\'s name:')
        c.drawString(start_2, y, "{} {}".format(emp_name, emp_surname))
        y -= spread

        c.drawString(start, y, 'Gross Salary :')
        c.drawString(start_2, y, "{}".format(emp_gross_salary))
        y -= spread

        c.drawString(start, y, 'Pension Contribution :')
        c.drawString(start_2, y, "{}".format(emp_pension_contribution))
        y -= spread

        c.drawString(start, y, 'Health Insurance :')
        c.drawString(start_2, y, "{}".format(emp_health_insurance))
        y -= spread

        c.drawString(start, y, 'Personal Income Tax :')
        c.drawString(start_2, y, "{}".format(emp_personal_income_tax))
        y -= spread

        c.drawString(start, y, 'Bonus Payment :')
        c.drawString(start_2, y, "{}".format(emp_bonus_payment))
        y -= spread

        c.drawString(start, y, 'Deduction :')
        c.drawString(start_2, y, "{}".format(emp_deduction))
        y -= spread

        c.drawString(start, y, 'Net Salary :')
        c.drawString(start_2, y, "{}".format(emp_net_salary))
        y -= spread * 3

        c.drawString(start, y, 'Signature:')
        c.drawString(start_2, y, "________________________")
        y -= spread

        c.save()

create_paylisp()